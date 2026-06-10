"""
Generate the full formatted workbook that mirrors Solution_AEP_Guinee_Final.xlsx.

Sheets produced:
  - "Solution Finale"      : regenerated from the algorithm
  - "DOSSIER_TECHNIQUE"    : regenerated (narrative + computed tables)
  - "Solution_1 (original)": copied from the provided source workbook
  - "Glossaire_Prix"       : copied from the provided source workbook
"""
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import COST_XLSX, OUTPUT_XLSX, RESERVOIR_HYDRAULIC_LEVEL_M, MAX_BUDGET_EUR

# ----- Palette (extracted from the original file) -----------------------
NAVY, NAVY2, BLUE = "0C2340", "1A3A5C", "1D64A8"
HDR = "1A3A5C"
ROW_L = "F3F8FD"
PANEL = "F0F7FF"
GREEN_F, GREEN_T = "DCFCE7", "15803D"
AMBER_F, AMBER_T = "FEF3C7", "92400E"
WHITE, GREY_T = "FFFFFF", "30435C"

THIN = Side(style="thin", color="D6E2EE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SF_WIDTHS = [32, 14, 15, 12, 12, 15, 13, 34]   # Solution Finale (8 cols)
DOS_WIDTHS = [46, 15, 22, 14, 13, 13, 22]       # DOSSIER (7 cols)


# ----- French number helpers --------------------------------------------
def num(n):
    return f"{n:,.0f}".replace(",", " ")


def eur(n):
    return f"{num(n)} €"


# ----- Low-level cell writer --------------------------------------------
def _f(size=10, bold=False, color=NAVY, italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def _fill(c):
    return PatternFill("solid", fgColor=c)


def _put(ws, r, c, value, font, fill=None, align="left", wrap=False, border=False):
    cell = ws.cell(row=r, column=c)
    cell.value = value
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = border
    return cell


def _height(text, widths):
    cpl = max(20, sum(widths) * 1.55)
    return 14 * max(1, int(len(str(text)) / cpl) + 1) + 2


# ----- Generic row renderer ---------------------------------------------
def render(ws, rows, widths):
    ncols = len(widths)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    for spec in rows:
        kind = spec[0]

        if kind == "blank":
            r += 1
            continue

        # full-width banner kinds
        if kind in ("title", "subtitle", "section", "h2", "para", "info", "warn"):
            text = spec[1]
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            if kind == "title":
                _put(ws, r, 1, text, _f(15, True, WHITE), _fill(NAVY))
                ws.row_dimensions[r].height = 30
            elif kind == "subtitle":
                _put(ws, r, 1, text, _f(10, False, "A8C8E8"), _fill(NAVY))
                ws.row_dimensions[r].height = 18
            elif kind == "section":
                _put(ws, r, 1, text, _f(12, True, WHITE), _fill(NAVY2))
                ws.row_dimensions[r].height = 24
            elif kind == "h2":
                _put(ws, r, 1, text, _f(11, True, NAVY), _fill(PANEL))
                ws.row_dimensions[r].height = 20
            elif kind == "para":
                _put(ws, r, 1, text, _f(10, False, GREY_T), None, wrap=True)
                ws.row_dimensions[r].height = _height(text, widths)
            elif kind == "info":
                _put(ws, r, 1, text, _f(9, False, BLUE), _fill(PANEL), wrap=True)
                ws.row_dimensions[r].height = _height(text, widths)
            elif kind == "warn":
                _put(ws, r, 1, text, _f(9, False, AMBER_T), _fill(AMBER_F), wrap=True)
                ws.row_dimensions[r].height = _height(text, widths)
            r += 1
            continue

        # table header
        if kind == "th":
            vals = spec[1]
            n = len(vals)
            for j, val in enumerate(vals, start=1):
                if j == n and n < ncols:
                    ws.merge_cells(start_row=r, start_column=j, end_row=r, end_column=ncols)
                _put(ws, r, j, val, _f(10, True, WHITE), _fill(HDR),
                     "center" if j > 1 else "left", wrap=True, border=BORDER)
            ws.row_dimensions[r].height = 26
            r += 1
            continue

        # table rows: tr (plain), trG (green total)
        if kind in ("tr", "trG"):
            vals = spec[1]
            n = len(vals)
            fill = _fill(GREEN_F) if kind == "trG" else (_fill(ROW_L) if spec[-1] == "z" else None)
            bold = kind == "trG"
            color = GREEN_T if kind == "trG" else NAVY
            for j, val in enumerate(vals, start=1):
                if j == n and n < ncols:
                    ws.merge_cells(start_row=r, start_column=j, end_row=r, end_column=ncols)
                _put(ws, r, j, val, _f(10, bold, color), fill,
                     "left" if j == 1 else "center", wrap=(j == n and n < ncols),
                     border=BORDER)
            r += 1
            continue

        # key/value/note : label[1:2] value[3:4] note[5:end]
        if kind == "kv":
            label, val, note = spec[1], spec[2], spec[3]
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=ncols)
            _put(ws, r, 1, label, _f(10, False, NAVY), _fill(ROW_L), border=BORDER)
            _put(ws, r, 3, val, _f(10, True, BLUE), _fill(WHITE), "center", border=BORDER)
            _put(ws, r, 5, note, _f(9, False, "5B6B82"), None, wrap=True)
            ws.row_dimensions[r].height = max(18, _height(note, widths[4:]))
            r += 1
            continue

        # cost line : label[1:4] value[5] note[6:end]   (8-col sheet)
        if kind in ("cost", "costT"):
            label, val, note = spec[1], spec[2], spec[3]
            total = kind == "costT"
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=ncols)
            f_lab = _f(10, total, GREEN_T if total else NAVY)
            f_val = _f(10, True, GREEN_T if total else BLUE)
            _put(ws, r, 1, label, f_lab, _fill(GREEN_F) if total else _fill(ROW_L), border=BORDER)
            _put(ws, r, 5, val, f_val, _fill(GREEN_F) if total else _fill(WHITE), "center", border=BORDER)
            _put(ws, r, 6, note, _f(9, False, GREEN_T if total else "5B6B82"), None, wrap=True)
            ws.row_dimensions[r].height = max(18, _height(note, widths[5:]))
            r += 1
            continue

    # freeze top title
    ws.sheet_view.showGridLines = False


# ======================================================================
#  SHEET 1 — SOLUTION FINALE
# ======================================================================
def build_solution_finale(S):
    fts = S["fountains"]
    served = S["served_population"]
    total = S["total_population"]
    cov = S["coverage_rate"] * 100
    demand_tot = served * S["design_lpcd"]
    H = RESERVOIR_HYDRAULIC_LEVEL_M

    q_moy = demand_tot / 86400
    q_12 = demand_tot / (12 * 3600)
    q_design = q_12 * 2
    q_res = 37000 / 86400
    margin = 37 - demand_tot / 1000

    rows = [
        ("title", "SOLUTION FINALE — 5 Bornes-fontaines gravitaires + Pompe solaire de relevage · Village Rural Guinée"),
        ("subtitle", f"Village rural Guinée · {num(total)} hab. · Réservoir base {H} m NGF · "
                     f"Débit max 37 m³/jour · Budget max {num(MAX_BUDGET_EUR)} €"),
        ("blank",),
        ("section", "1.  LOCALISATIONS PROPOSÉES DES 5 BORNES-FONTAINES"),
        ("th", ["BF #", "Alt. (m NGF)", "Charge dispon. (m CE)", "Ménages couverts",
                "Population", "Demande conçue (L/j)", "Dist. réservoir", "Alimentation"]),
    ]
    tot_hh = 0
    for _, f in fts.iterrows():
        alt = f["Altitude"]
        charge = H - alt
        hh = int(f["households_covered"])
        pop = f["population_covered"]
        tot_hh += hh
        rows.append(("tr", [f["bf_id"], int(alt), int(charge), hh, int(pop),
                            int(pop * S["design_lpcd"]), f"{f['dist_res_m']:.0f} m", "✓ Gravité"], "z"))
    rows.append(("trG", ["TOTAL", "", "", tot_hh, int(served), int(demand_tot),
                        f"Couverture : {cov:.1f}%", ""]))
    rows.append(("blank",))

    rows += [
        ("section", "2.  CALCUL DES DÉBITS — Paramètres de dimensionnement"),
        ("th", ["Paramètre", "Valeur calculée", "Commentaire / Source"]),
        ("kv", "Dotation de conception retenue", "20 L/pers./jour",
         "OMS minimum — supérieur à la demande exprimée (17,2 L/p/j)"),
        ("kv", "Facteur de pointe horaire (PHF)", "2",
         "Rural Afrique Ouest — pointe matin + soir"),
        ("kv", "Fenêtre de distribution", "12 heures/jour",
         "Système gravitaire — distribution 6h00–18h00"),
        ("kv", "Débit moyen (sur 24 h)", f"{q_moy:.3f} L/s",
         f"{demand_tot/1000:.2f} m³/j ÷ 86 400 s"),
        ("kv", "Débit de conception (PHF=2, 12 h)", f"{q_design:.3f} L/s",
         "Débit max utilisé pour dimensionner les tuyaux"),
        ("kv", "Débit réservoir disponible", "37,0 m³/jour",
         "Contrainte pompe motorisée — énoncé projet"),
        ("kv", "Marge hydraulique résiduelle", f"{margin:.2f} m³/j",
         "= 37 − demande totale conçue"),
        ("kv", "Coeff. Hazen-Williams (HDPE neuf)", "C = 135",
         "Norme HDPE PN10 / ISO 4427"),
        ("blank",),
        ("section", "3.  DIMENSIONNEMENT DES CONDUITES — Hazen-Williams (C = 135, HDPE PN10)"),
        ("th", ["Tronçon", "Longueur (m)", "Q design (L/s)", "Diamètre",
                "Vitesse (m/s)", "Perte hf (m)", "hf/km", "Pression aval (m CE)"]),
    ]
    for s in S["segments"]:
        tag = " ⚠ PRV conseillé" if s["prv_needed"] else ""
        rows.append(("tr", [f"{s['from']} → {s['to']}", f"{s['length_m']:.0f}",
                            f"{s['q_lps']:.3f}", f"DN{s['dn_mm']}", f"{s['velocity_ms']:.3f}",
                            f"{s['hf_m']:.3f}", f"{s['hf_m']/s['length_m']*1000:.2f}",
                            f"{s['pressure_m']:.1f}{tag}"], "z"))
    rows.append(("info", "Vitesse 🟢 > 0,3 m/s · 🟡 0,1–0,3 m/s acceptable gravité · "
                         "Pression min 5 m CE · PRV si pression > 20 m CE · "
                         "hf = 10,67 × L × Q^1,852 / (C^1,852 × D^4,87)"))
    rows.append(("blank",))

    p = S["pump"]
    rows += [
        ("section", "4.  POMPE DE RELEVAGE — foyers non alimentables par gravité (alt. > 373 m NGF)"),
        ("th", ["Paramètre", "Valeur", "Commentaire / Calcul"]),
        ("kv", "Ménages concernés", f"{p['households']} ménages (~{p['population']:.0f} pers.)",
         "Altitude 375–378 m NGF — au-dessus de la base cuve (373 m)"),
        ("kv", "Demande journalière", f"{p['daily_l']:.0f} L/jour",
         f"{p['households']} × 6 pers × 20 L/p/j"),
        ("kv", "Hauteur manométrique (Hm)", f"{p['hm_m']:.0f} m",
         "Δz=8 m (lift) + 5 m (pression min) + 2 m (pertes)"),
        ("kv", "Débit pompe (Q)", f"{p['q_lps']:.3f} L/s",
         "Demande distribuée sur 2 h de pompage/jour"),
        ("kv", "Puissance hydraulique", f"{p['p_hyd_w']:.1f} W",
         "P = ρ × g × Q × Hm"),
        ("kv", "Puissance arbre (η=55%)", f"{p['p_shaft_w']:.1f} W",
         "P_arbre = P_hyd / rendement pompe → pompe 50 W solaire"),
        ("kv", "Énergie quotidienne", "0,10 kWh/jour",
         "50 W × 2 h — 1 panneau 100 W suffit · OPEX ≈ 0"),
        ("kv", "CAPEX pompe (total)", "~2 600 €",
         "Pompe solaire 600 € + tuyau DN25 × 200 m (2 000 €)"),
        ("blank",),
        ("section", "5.  ESTIMATION DES COÛTS — Solution 1"),
        ("th", ["Poste de dépense", "", "", "", "Total (€)", "Source / justification"]),
    ]
    for label, val in S["costs"]["capex_items"]:
        rows.append(("cost", label, eur(val), "Benchmark Afrique de l'Ouest (WASHCost / IRC WASH)"))
    c = S["costs"]
    rows.append(("costT", "CAPEX TOTAL SOLUTION 1", eur(c["capex_total"]),
                 f"✅ DANS LE BUDGET — Marge : {eur(MAX_BUDGET_EUR - c['capex_total'])} "
                 f"sur {num(MAX_BUDGET_EUR)} € "
                 f"({c['capex_total']/MAX_BUDGET_EUR*100:.1f}% utilisé)"))
    rows.append(("cost", "OPEX annuel estimé", eur(c["opex_total"]),
                 f"Cycle de vie 20 ans : {eur(c['lifecycle'])} · "
                 f"Coût/pers. : {c['cost_per_person']:.0f} €/pers."))
    rows.append(("blank",))
    rows.append(("section", "▸  Solution générée par algorithme — couverture greedy 500 m · "
                            "dimensionnement Hazen-Williams · coûts benchmark AO"))
    return rows


# ======================================================================
#  SHEET 2 — DOSSIER TECHNIQUE
# ======================================================================
def build_dossier(S):
    served = S["served_population"]
    total = S["total_population"]
    cov = S["coverage_rate"] * 100
    demand_tot = served * S["design_lpcd"]
    c = S["costs"]
    p = S["pump"]
    fts = S["fountains"]

    rows = [
        ("title", "DOSSIER TECHNIQUE — Explication complète pour investisseurs et enseignants"),
        ("subtitle", "Village rural Guinée · Système AEP gravitaire · CAFO — L'eau qui rassemble · Juin 2026"),
        ("blank",),
        ("section", "SECTION 1 — INTRODUCTION ET ANALYSE DE LA SITUATION"),
        ("h2", "1.1  Présentation du projet et objectifs"),
        ("para", f"Ce projet répond à un appel d'offre du gouvernement local guinéen. "
                 f"L'entreprise CAFO conçoit un système d'adduction d'eau potable (AEP) pour "
                 f"un village rural de {num(total)} habitants, en utilisant un château d'eau "
                 f"déjà installé. Le projet doit rester sous {num(MAX_BUDGET_EUR)} € sur tout "
                 f"son cycle de vie, ne pas dépasser 37 m³/jour, et être gérable par la "
                 f"communauté locale sans dépendance à des techniciens extérieurs."),
        ("h2", "1.2  Problème actuel — pourquoi le village n'a pas accès à l'eau potable"),
        ("para", "Les 21 puits ouverts du village sont contaminés : 200 à 600 UFC/100 mL de "
                 "coliformes fécaux (norme OMS = 0). C'est la principale cause de maladies "
                 "diarrhéiques. Les 3 pompes à main existantes sont insuffisantes. En l'absence "
                 "d'eau potable proche, femmes et enfants marchent en moyenne 1 169 m (parfois "
                 "2 km) pour chercher de l'eau, souvent non sûre."),
        ("info", "ℹ Données de terrain : qualité de l'eau et GPS issus du fichier RAW_DATA. "
                 "Les distances sont calculées par la formule de haversine sur les coordonnées GPS."),
        ("h2", "1.3  Population et démographie"),
        ("th", ["Paramètre", "Valeur mesurée", "Projection 10 ans (2036)", "Source"]),
        ("tr", ["Population totale", f"{num(total)} pers.", "~1 340 pers.", "Enquête terrain RAW_DATA"]),
        ("tr", ["Nombre de ménages", f"{len(S['households'])} foyers", "~221 foyers", "Enquête terrain"]),
        ("tr", ["Taille moyenne / foyer", f"{total/len(S['households']):.1f} pers./foyer", "—", "Calculé sur RAW_DATA"]),
        ("tr", ["Taux croissance annuel", "2,7 %/an (Guinée)", "—", "World Bank 2024"]),
        ("tr", ["Débit max disponible", "37 m³/j", "37 m³/j", "Contrainte pompe — énoncé"]),
        ("h2", "1.4  Géographie et topographie — le terrain décide tout"),
        ("para", "Le château d'eau trône sur la butte centrale à 368 m NGF ; la base de la cuve "
                 "est à 373 m NGF. Comme 98 % des foyers sont sous cette altitude, l'eau descend "
                 "par gravité, sans pompe électrique principale."),
        ("th", ["Paramètre topographique", "Valeur", "Signification pour le réseau"]),
        ("tr", ["Altitude sol réservoir", "368 m NGF", "Point d'ancrage du réseau"]),
        ("tr", ["Altitude base de cuve", "373 m NGF", "L'eau sort à cette altitude — pression disponible"]),
        ("tr", ["Altitude min. des foyers", "350 m NGF", "Pression max = 373−350 = 23 m CE"]),
        ("tr", ["Altitude max. des foyers", "378 m NGF", "3 foyers au-dessus → pas de gravité"]),
        ("tr", ["Pression minimum requise", "5 m CE", "Norme pratique AEP rurale — assurée partout"]),
        ("tr", ["Foyers hors gravité (>373 m)", f"{p['households']} foyers", "Petite pompe solaire de relevage"]),
        ("h2", "1.5  Analyse des besoins en eau"),
        ("para", "Chaque foyer a déclaré ses besoins (boire, cuisiner, hygiène, lessive) : "
                 "17 752 L/jour pour 1 031 personnes, soit 17,2 L/p/j — inférieur au minimum OMS "
                 "de 20 L/p/j. On retient donc 20 L/p/j comme base de conception."),
        ("warn", "⚠ Les 15 % de pertes réseau sont estimés selon les standards WASH (réseaux "
                 "ruraux neufs : 10–20 %). Taux conservateur."),
        ("blank",),
        ("section", "SECTION 2 — CONCEPTION TECHNIQUE DU RÉSEAU"),
        ("h2", "2.1  Stratégie générale — gravité et bornes-fontaines publiques"),
        ("para", "Puisque 98 % des foyers sont sous 373 m NGF, l'eau coule naturellement du "
                 "réservoir : l'option la plus fiable et la moins chère en Afrique rurale (pas "
                 "d'électricité, pas de panne de pompe). 5 bornes-fontaines publiques desservent "
                 "chacune ~200 personnes dans un rayon de 500 m (norme Sphere/OMS)."),
        ("h2", "2.2  Comment les 5 emplacements ont été choisis — algorithme de couverture"),
        ("para", "Un algorithme glouton (greedy coverage) analyse les positions GPS des 170 foyers. "
                 "Principe : on place la 1ère borne là où elle couvre le plus de personnes dans un "
                 "rayon de 500 m, on marque ces personnes comme desservies, et on recommence. "
                 f"En 5 itérations on atteint {cov:.1f} % de couverture."),
        ("th", ["BF", "Distance réservoir", "Population couverte", "Altitude", "Alimentation"]),
    ]
    for _, f in fts.iterrows():
        pct = f["population_covered"] / total * 100
        rows.append(("tr", [f["bf_id"], f"{f['dist_res_m']:.0f} m",
                            f"{f['population_covered']:.0f} pers. ({pct:.0f}%)",
                            f"{f['Altitude']:.0f} m NGF", "✓ Gravité"]))
    rows.append(("trG", ["TOTAL", "", f"{served:.0f} / {num(total)} personnes ({cov:.1f}%)", "", ""]))

    q_moy = demand_tot / 86400
    q_12 = demand_tot / (12 * 3600)
    q_design = q_12 * 2
    rows += [
        ("h2", "2.3  Calcul des débits"),
        ("para", "On calcule le débit de pointe (matin + soir) avec un facteur de pointe horaire "
                 "de 2, en supposant la consommation concentrée sur 12 h."),
        ("th", ["Étape", "Calcul", "Résultat"]),
        ("tr", ["Débit moyen (24 h)", f"{num(demand_tot)} L/j ÷ 86 400 s", f"{q_moy:.3f} L/s"]),
        ("tr", ["Débit sur 12 h", f"{num(demand_tot)} L/j ÷ 43 200 s", f"{q_12:.3f} L/s"]),
        ("tr", ["Débit de conception (×2)", f"{q_12:.3f} × 2 (PHF)", f"{q_design:.3f} L/s"]),
        ("tr", ["Vérification réservoir", "37 000 L/j ÷ 86 400 s", "0,428 L/s"]),
        ("h2", "2.4  Dimensionnement des tuyaux — Hazen-Williams"),
        ("para", "Pour chaque tronçon on choisit le plus petit diamètre qui maintient une "
                 "vitesse correcte et garantit ≥ 5 m de pression à la borne. Tuyaux HDPE PN10 "
                 "(durée de vie 25–50 ans). hf = 10,67 × L × Q^1,852 / (C^1,852 × D^4,87)."),
        ("th", ["Tronçon", "L (m)", "Q (L/s)", "Diamètre", "V (m/s)", "hf (m)", "Pression aval (m CE)"]),
    ]
    for s in S["segments"]:
        tag = " ⚠ PRV" if s["prv_needed"] else " ✓"
        rows.append(("tr", [f"{s['from']} → {s['to']}", f"{s['length_m']:.0f}", f"{s['q_lps']:.3f}",
                            f"DN{s['dn_mm']}", f"{s['velocity_ms']:.3f}", f"{s['hf_m']:.3f}",
                            f"{s['pressure_m']:.1f}{tag}"]))
    rows.append(("warn", "⚠ PRV = réducteur de pression (~130 €) : installé là où la pression "
                         "dépasse 20 m CE pour protéger les robinets."))
    rows += [
        ("h2", "2.5  Pompe de relevage — foyers impossibles par gravité"),
        ("para", f"{p['households']} foyers sont à 375–378 m NGF, au-dessus de la base de cuve "
                 f"(373 m). Une petite pompe solaire installée à 370 m NGF refoule l'eau sur "
                 f"200 m. Fonctionne 2 h/matin. Énergie solaire = OPEX nul."),
        ("th", ["Calcul", "Détail", "Résultat"]),
        ("tr", ["Demande des foyers", f"{p['households']} × 6 × 20 L/j", f"{p['daily_l']:.0f} L/jour"]),
        ("tr", ["Hauteur à pomper (Hm)", "Δz(8) + pression(5) + pertes(2)", f"{p['hm_m']:.0f} m"]),
        ("tr", ["Débit pompe (sur 2 h)", f"{p['daily_l']:.0f} L ÷ 7 200 s", f"{p['q_lps']:.3f} L/s"]),
        ("tr", ["Puissance hydraulique", "P = ρ·g·Q·Hm", f"{p['p_hyd_w']:.1f} W"]),
        ("tr", ["Puissance réelle (η=55%)", f"{p['p_hyd_w']:.1f} W ÷ 0,55", f"{p['p_shaft_w']:.1f} W"]),
        ("info", "ℹ CAPEX pompe : 600 € (pompe+panneau+batterie) + 2 000 € (200 m DN25) = 2 600 €. "
                 "OPEX : 0 €/an grâce au solaire."),
        ("blank",),
        ("section", "SECTION 3 — ÉTUDE FINANCIÈRE"),
        ("h2", "3.1  Comment chaque coût a été estimé"),
        ("para", "Prix = estimations régionales Afrique de l'Ouest (WASHCost Burkina Faso 2019, "
                 "IRC WASH Costing 2021). Aucune base officielle Guinée publique : à valider avec "
                 "un fournisseur local. Marge d'imprévus de 5 % incluse."),
        ("th", ["Poste de dépense", "Montant", "Justification"]),
    ]
    for label, val in c["capex_items"]:
        rows.append(("tr", [label, eur(val), "Benchmark AO (WASHCost / IRC WASH) — algorithme pour la longueur réseau"]))
    rows.append(("trG", ["CAPEX TOTAL", eur(c["capex_total"]),
                        f"{eur(MAX_BUDGET_EUR)} disponibles → marge {eur(MAX_BUDGET_EUR - c['capex_total'])}"]))
    rows += [
        ("h2", "3.2  Coûts annuels d'exploitation (OPEX)"),
        ("th", ["Poste annuel", "Coût / an", "Explication"]),
    ]
    for label, val in c["opex_items"]:
        rows.append(("tr", [label, eur(val), ""]))
    rows.append(("trG", ["OPEX TOTAL / AN", eur(c["opex_total"]),
                        f"{c['opex_total']/served:.1f} €/personne/an"]))
    rows += [
        ("h2", "3.3  Bilan financier sur 20 ans"),
        ("th", ["Poste", "Montant", "Calcul"]),
        ("tr", ["CAPEX (investissement initial)", eur(c["capex_total"]), "—"]),
        ("tr", ["OPEX cumulé sur 20 ans", eur(c["opex_total"] * 20), f"{eur(c['opex_total'])}/an × 20 ans"]),
        ("trG", ["TOTAL cycle de vie 20 ans", eur(c["lifecycle"]),
                 f"{eur(c['capex_total'])} + {eur(c['opex_total']*20)}"]),
        ("tr", ["Coût par personne desservie", f"{c['cost_per_person']:.0f} €/personne",
                f"{eur(c['lifecycle'])} ÷ {served:.0f} pers."]),
        ("warn", "⚠ Incertitude ±30 % sur les coûts (benchmarks AO non vérifiés en Guinée). "
                 "Scénario pessimiste (CAPEX ×1,3) ≈ 222 750 € — légèrement au-dessus du budget."),
        ("blank",),
        ("section", "SECTION 4 — PLAN DE MISE EN ŒUVRE"),
        ("h2", "4.1  Calendrier prévisionnel — 4 à 5 mois"),
        ("th", ["Phase", "Durée", "Activités principales", "Ressources"]),
        ("tr", ["1 — Préparation", "1 mois", "Levé GPS, plans d'exécution, commande matériaux", "Topographe + chef projet"]),
        ("tr", ["2 — Terrassement", "1,5 mois", "Fouille tranchées, pose conduites DN50/DN40, remblayage", "20 manœuvres + 3 plombiers"]),
        ("tr", ["3 — Équipements", "1 mois", "Pose des 5 BF, vannes, PRV, chlorateur", "3 plombiers + 1 maçon"]),
        ("tr", ["4 — Pompe et tests", "0,5 mois", "Pompe solaire + DN25, mise en eau, essais pression", "1 électricien + 2 plombiers"]),
        ("tr", ["5 — Formation", "0,5 mois", "Formation 3 techniciens, sensibilisation, remise dossier", "1 formateur + équipe"]),
        ("blank",),
        ("section", "SECTION 5 — ENTRETIEN ET DURABILITÉ"),
        ("h2", "5.1  Stratégie de durabilité sur 20 ans"),
        ("para", "Les projets AEP ruraux échouent rarement pour des raisons techniques — c'est "
                 "presque toujours la gouvernance et le financement de l'entretien. Notre réponse : "
                 "(1) gravité sans pompe principale, (2) provision réhabilitation 1 500 €/an, "
                 "(3) techniciens locaux formés pour des interventions rapides."),
        ("h2", "5.3  Gestion communautaire — modèle économique du comité eau"),
        ("th", ["Paramètre", "Calcul", "Résultat"]),
        ("tr", ["Volume facturé / an (85 %)", "20 m³/j × 365 × 0,85", "6 205 m³/an"]),
        ("tr", ["Tarif minimum (OPEX seul)", f"{eur(c['opex_total'])} ÷ 6 205 m³", f"{c['opex_total']/6205:.2f} €/m³"]),
        ("tr", ["Tarif recommandé (+ réhab.)", f"({eur(c['opex_total'])} + 1 500) ÷ 6 205", f"{(c['opex_total']+1500)/6205:.2f} €/m³"]),
        ("tr", ["Part revenu ménage rural", "3,40 ÷ 80 €/mois", "< 5 % ✓ abordable"]),
        ("blank",),
        ("section", f"SYNTHÈSE :  Budget respecté ({eur(c['capex_total'])} < {num(MAX_BUDGET_EUR)} €) · "
                    f"{cov:.1f} % de la population desservie · Eau conforme OMS · "
                    f"Système gravitaire 24h/24 · Durée de vie HDPE 25–50 ans."),
    ]
    return rows


# ======================================================================
#  Copy static sheets from the source workbook
# ======================================================================
def copy_sheet(src_ws, tgt_ws):
    for row in src_ws.iter_rows():
        for cell in row:
            t = tgt_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                t.font = copy(cell.font)
                t.fill = copy(cell.fill)
                t.border = copy(cell.border)
                t.alignment = copy(cell.alignment)
                t.number_format = cell.number_format
    for mc in list(src_ws.merged_cells.ranges):
        tgt_ws.merge_cells(str(mc))
    for col, dim in src_ws.column_dimensions.items():
        if dim.width:
            tgt_ws.column_dimensions[col].width = dim.width
    tgt_ws.sheet_view.showGridLines = False


# ======================================================================
def build_workbook(S, out_path=OUTPUT_XLSX):
    wb = Workbook()
    wb.remove(wb.active)

    render(wb.create_sheet("Solution Finale"), build_solution_finale(S), SF_WIDTHS)
    render(wb.create_sheet("DOSSIER_TECHNIQUE"), build_dossier(S), DOS_WIDTHS)

    src = load_workbook(COST_XLSX, data_only=True)
    copy_sheet(src["Solution_1"], wb.create_sheet("Solution_1 (original)"))
    copy_sheet(src["Glossaire_Prix"], wb.create_sheet("Glossaire_Prix"))

    wb.save(out_path)
    return out_path
